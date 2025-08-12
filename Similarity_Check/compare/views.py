from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, FileResponse, Http404
from .models import Book
from .forms import BookForm
import PyPDF2
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import os
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.utils.encoding import smart_str
from django.contrib.auth.models import User
from .forms import CustomUserCreationForm
from django.contrib.auth.models import User
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import Group
from .models import ComparisonHistory
from .utils import extract_file_content, calculate_similarity
import os
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Book, ComparisonHistory
from .utils import extract_file_content, calculate_similarity
from django.contrib import messages
from django.db.models import Count
from .models import ComparisonHistory
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.contrib.auth.models import User
from .models import Book, ComparisonHistory
from .utils import extract_file_content, calculate_similarity
from .forms import EditUserWithPasswordForm
import openpyxl
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.contrib.auth.models import Group
from django.contrib.auth import get_user_model
import openpyxl
from django.contrib.auth.models import User, Group
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils.decorators import method_decorator
from django.views import View
from django.core.exceptions import ValidationError
from django.contrib.auth.hashers import make_password
from django.shortcuts import render
import os

User = get_user_model()

@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_superuser)
def preview_user_import(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        path = default_storage.save('temp_import_users.xlsx', file)
        try:
            wb = openpyxl.load_workbook(default_storage.path(path))
            sheet = wb.active

            expected_columns = ['Username', 'Email', 'Password', 'Role']
            first_row = [cell.value for cell in sheet[1]]

            if first_row != expected_columns:
                return JsonResponse({'status': 'error', 'message': 'Invalid column headers.'})

            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(row):
                    data.append({
                        'username': row[0],
                        'email': row[1],
                        'password': row[2],
                        'role': row[3]
                    })

            return JsonResponse({'status': 'ok', 'data': data})
        finally:
            # Clean up the temporary file
            os.remove(default_storage.path(path))
    
    return JsonResponse({'status': 'error', 'message': 'No file uploaded.'})

@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_superuser)
def upload_user_import(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        path = default_storage.save('temp_import_users.xlsx', file)
        try:
            wb = openpyxl.load_workbook(default_storage.path(path))
            sheet = wb.active

            expected_columns = ['Username', 'Email', 'Password', 'Role']
            first_row = [cell.value for cell in sheet[1]]

            if first_row != expected_columns:
                return JsonResponse({'status': 'error', 'message': 'Invalid column headers.'})

            imported = []
            skipped = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                username, email, password, role = row

                if not username or not email or not password or not role:
                    skipped.append(username or "Unknown")
                    continue

                if User.objects.filter(username=username).exists():
                    skipped.append(username)
                    continue

                user = User.objects.create(
                    username=username,
                    email=email,
                    password=make_password(password),
                    is_superuser=(role == 'Admin'),
                    is_staff=(role == 'Admin')
                )

                group, _ = Group.objects.get_or_create(name=role)
                user.groups.add(group)

                imported.append(username)

            return JsonResponse({
                'status': 'ok',
                'imported': imported,
                'skipped': skipped,
                'message': 'Users uploaded successfully.'
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
        finally:
            os.remove(default_storage.path(path))
    
    return JsonResponse({'status': 'error', 'message': 'No file uploaded or invalid request method.'})

@method_decorator(csrf_exempt, name='dispatch')
class ImportUsersView(View):
    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return JsonResponse({'status': False, 'message': 'No file uploaded'}, status=400)

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Expected header
            expected_columns = ['username', 'email', 'password', 'role']
            header = [cell.value.lower() if cell.value else '' for cell in sheet[1]]

            if header != expected_columns:
                return JsonResponse({'status': False, 'message': 'Invalid Excel format. Columns must be: username, email, password, role'}, status=400)

            imported = []
            skipped = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                username, email, password, role = row

                if not username or not email or not password or not role:
                    skipped.append(username or "Unknown")
                    continue

                if User.objects.filter(username=username).exists():
                    skipped.append(username)
                    continue

                user = User.objects.create(
                    username=username,
                    email=email,
                    password=make_password(password),
                    is_superuser=(role == 'Admin'),
                    is_staff=(role == 'Admin')
                )

                group, _ = Group.objects.get_or_create(name=role)
                user.groups.add(group)

                imported.append(username)

            return JsonResponse({
                'status': True,
                'imported': imported,
                'skipped': skipped
            })

        except Exception as e:
            return JsonResponse({'status': False, 'message': str(e)}, status=500)

def logout_view(request):
    logout(request)
    return redirect('login')

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password")
    return render(request, 'compare/login.html')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def register_user(request):
    form = CustomUserCreationForm(request.POST or None)

    if request.method == 'POST':
        role = request.POST.get('role')

        if form.is_valid():
            user = form.save(commit=False)

            if role == 'Admin':
                user.is_superuser = True
                user.is_staff = True
            else:
                user.is_superuser = False
                user.is_staff = False

            user.save()

            if role:
                group, created = Group.objects.get_or_create(name=role)
                user.groups.add(group)

            messages.success(request, "User created successfully.")
            return redirect('dashboard')

    role = get_user_role(request.user)

    return render(request, 'compare/register_user.html', {
        'form': form,
        'role': role
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def edit_user(request, user_id):
    user_obj = get_object_or_404(User, pk=user_id)
    form = EditUserWithPasswordForm(request.POST or None, instance=user_obj)
    user_groups = user_obj.groups.values_list('name', flat=True)
    selected_role = user_groups[0] if user_groups else 'Student'

    if request.method == 'POST' and form.is_valid():
        user = form.save(commit=False)
        password = form.cleaned_data.get("password1")
        if password:
            user.set_password(password)
        else:
            user.password = user_obj.password

        role = request.POST.get('role')
        user.groups.clear()
        if role:
            group, _ = Group.objects.get_or_create(name=role)
            user.groups.add(group)

        user.is_superuser = (role == 'Admin')
        user.is_staff = (role == 'Admin')
        user.save()

        messages.success(request, 'User updated successfully.')
        return redirect('users_list')

    # 👉 add this
    current_role = get_user_role(request.user)

    return render(request, 'compare/edit_user.html', {
        'form': form,
        'edited_user': user_obj,      # renamed (no longer overrides {{ user }})
        'selected_role': selected_role,
        'role': current_role          # 👉 pass role to base_dashboard
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_user(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'User deleted successfully.')
        return redirect('users_list')
    return redirect('users_list')

from django.shortcuts import render
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
import json

@login_required
@user_passes_test(lambda u: u.is_superuser)
def users_list(request):
    users = User.objects.all()
    role = get_user_role(request.user)
    
    # Serialize users data for JavaScript
    users_data = [
        {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'date_joined': user.date_joined.strftime('%Y-%m-%d'),
            'is_superuser': user.is_superuser,
            'groups': [group.name for group in user.groups.all()]
        } for user in users
    ]
    users_json = json.dumps(users_data)

    return render(request, 'compare/users_list.html', {
        'users': users,
        'users_json': users_json,
        'role': role
    })

def get_user_role(user):
    if user.is_superuser or user.groups.filter(name='Admin').exists():
        return 'Admin'
    elif user.groups.filter(name='Supervisor').exists():
        return 'Supervisor'
    else:
        return 'Student'

def read_pdf(file):
    try:
        reader = PyPDF2.PdfReader(file)
        text = ""
        for page in reader.pages:
            extracted_text = page.extract_text()
            if extracted_text:
                text += extracted_text
        return text
    except Exception as e:
        return f"Error extracting PDF content: {str(e)}"

from docx import Document as DocxDocument

def read_docx(file):
    try:
        doc = DocxDocument(file)
        text = "\n".join([para.text for para in doc.paragraphs])
        return text
    except Exception as e:
        return f"Error extracting DOCX content: {str(e)}"

def extract_file_content(file):
    ext = file.name.split('.')[-1].lower()
    if ext == 'pdf':
        return read_pdf(file)
    elif ext in ['docx', 'doc']:
        return read_docx(file)
    else:
        return None


def calculate_similarity(text1, text2):
    try:
        tfidf_vectorizer = TfidfVectorizer(stop_words='english')
        tfidf_matrix = tfidf_vectorizer.fit_transform([text1, text2])
        similarity_matrix = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])
        return similarity_matrix[0][0]
    except Exception as e:
        print(f"Error calculating similarity: {str(e)}")
        return 0.0

@login_required
def check_title(request):
    message = ''
    if request.method == 'POST':
        title = request.POST.get('title')
        normalized_title = title.strip().lower()
        books = Book.objects.filter(book_title__icontains=title)
        matched_books = [book for book in books if book.book_title.strip().lower() == normalized_title]
        if matched_books:
            matched_book = matched_books[0]
            message = f"Your title '{title}' is already covered in the book '{matched_book.book_title}' by {matched_book.authors}."
        else:
            message = f"No books found matching the title '{title}'."

        ComparisonHistory.objects.create(
            user=request.user,
            compared_with=None,
            similarity_score=0.0,
            uploaded_title=title
        )

        return render(request, 'compare/check_title.html', {'message': message})

    return render(request, 'compare/check_title.html')

def view_book_file(request, book_id):
    try:
        book = Book.objects.get(book_id=book_id)
        file_path = book.file.path
        ext = book.file.name.split('.')[-1].lower()

        if ext == 'pdf':
            response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{smart_str(book.file.name)}"'
            return response
        elif ext in ['doc', 'docx']:
            response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
            response['Content-Disposition'] = f'attachment; filename="{smart_str(book.file.name)}"'
            return response
        else:
            raise Http404("Unsupported file type.")
    except Book.DoesNotExist:
        raise Http404("Book not found.")
    except Exception as e:
        raise Http404(f"Error loading file: {e}")

@login_required
def compare_uploaded_books(request):
    if request.method == 'POST' and request.FILES.get('file1'):
        file1 = request.FILES['file1']
        uploaded_content = extract_file_content(file1)

        if not uploaded_content:
            return render(request, 'compare/compare_result.html', {
                'message': 'Failed to extract content. Unsupported file type or empty content.'
            })

        file_name = os.path.splitext(file1.name)[0].strip().lower()
        books = Book.objects.all()

        matched_book = None
        for book in books:
            if file_name == book.book_title.strip().lower():
                matched_book = book
                break

        if matched_book:
            matched_content = extract_file_content(matched_book.file)
            if matched_content:
                similarity = calculate_similarity(uploaded_content, matched_content)
                rounded_score = round(similarity * 100, 2)
                message = "Your book is eligible."
                if similarity > 0.20:
                    message = "Your book is not eligible due to high similarity."

                ComparisonHistory.objects.create(
                    user=request.user,
                    compared_with=matched_book,
                    similarity_score=similarity,
                    uploaded_title=file_name
                )

                return render(request, 'compare/compare_result.html', {
                    'most_similar_book': matched_book,
                    'similarity_percentage': rounded_score,
                    'eligibility_message': message
                })

        highest_similarity = 0
        most_similar_book = None
        for book in books:
            if book.file:
                book_content = extract_file_content(book.file)
                if book_content:
                    similarity = calculate_similarity(uploaded_content, book_content)
                    if similarity > highest_similarity:
                        highest_similarity = similarity
                        most_similar_book = book

        rounded_score = round(highest_similarity * 100, 2)
        message = "Your book is eligible."
        if highest_similarity > 0.20:
            message = "Your book is not eligible due to high similarity."

        if most_similar_book:
            ComparisonHistory.objects.create(
                user=request.user,
                compared_with=most_similar_book,
                similarity_score=highest_similarity,
                uploaded_title=file_name
            )

        return render(request, 'compare/compare_result.html', {
            'most_similar_book': most_similar_book,
            'similarity_percentage': rounded_score,
            'eligibility_message': message
        })

    return redirect('index')

from django.db.models import Q
from django.contrib.auth.models import Group

@login_required
def dashboard(request):
    user = request.user
    role = get_user_role(user)

    total_books = Book.objects.count()

    if role == 'Admin':
        total_users = User.objects.count()
        total_admins = User.objects.filter(is_superuser=True).count()
        total_supervisors = User.objects.filter(groups__name='Supervisor').count()
        total_students = total_users - total_admins - total_supervisors

        recent_comparisons = ComparisonHistory.objects.select_related('compared_with', 'user').order_by('-timestamp')[:10]
        title_checks = ComparisonHistory.objects.filter(compared_with__isnull=True)
        document_checks = ComparisonHistory.objects.filter(compared_with__isnull=False)

    elif role == 'Supervisor':
        total_users = total_admins = 0
        total_supervisors = User.objects.filter(groups__name='Supervisor').count()
        student_group = Group.objects.get(name='Student')
        student_users = User.objects.filter(groups=student_group)
        total_students = student_users.count()

        recent_comparisons = ComparisonHistory.objects.filter(user__in=student_users).select_related('compared_with').order_by('-timestamp')[:10]
        title_checks = ComparisonHistory.objects.filter(user__in=student_users, compared_with__isnull=True)
        document_checks = ComparisonHistory.objects.filter(user__in=student_users, compared_with__isnull=False)

    else:
        total_users = total_admins = 0
        total_supervisors = User.objects.filter(groups__name='Supervisor').count()
        total_students = User.objects.filter(groups__name='Student').count()

        recent_comparisons = ComparisonHistory.objects.filter(user=user).select_related('compared_with').order_by('-timestamp')[:10]
        title_checks = ComparisonHistory.objects.filter(user=user, compared_with__isnull=True)
        document_checks = ComparisonHistory.objects.filter(user=user, compared_with__isnull=False)

    title_checks_count = title_checks.count()
    document_checks_count = document_checks.count()

    ineligible_document_checks = document_checks.filter(similarity_score__gt=0.20).count()
    eligible_document_checks = document_checks_count - ineligible_document_checks
    document_accuracy = (eligible_document_checks / document_checks_count) * 100 if document_checks_count > 0 else 0

    eligible_title_count = 0
    for check in title_checks:
        normalized = check.uploaded_title.strip().lower()
        if Book.objects.filter(book_title__iexact=normalized).exists():
            eligible_title_count += 1
    not_eligible_title_count = title_checks_count - eligible_title_count
    title_accuracy = (eligible_title_count / title_checks_count) * 100 if title_checks_count > 0 else 0

    context = {
        'role': role,
        'total_users': total_users,
        'total_admins': total_admins,
        'total_supervisors': total_supervisors,
        'total_students': total_students,
        'total_books': total_books,
        'recent_comparisons': recent_comparisons,
        'title_checks_count': title_checks_count,
        'document_checks_count': document_checks_count,
        'document_accuracy': round(document_accuracy, 2),
        'title_accuracy': round(title_accuracy, 2),
        'eligible_doc_count': eligible_document_checks,
        'not_eligible_doc_count': ineligible_document_checks,
        'eligible_title_count': eligible_title_count,
        'not_eligible_title_count': not_eligible_title_count,
    }

    return render(request, 'compare/dashboard.html', context)

def index(request):
    return render(request, 'compare/index.html')

def books_list(request):
    books = Book.objects.all()
    return render(request, 'compare/book_list.html', {'books': books})

def add_or_edit_book(request, book_id=None):
    book = get_object_or_404(Book, book_id=book_id) if book_id else None
    form = BookForm(request.POST or None, request.FILES or None, instance=book)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('books_list')
    return render(request, 'compare/registration.html', {'form': form})

def delete_book(request, book_id):
    book = get_object_or_404(Book, book_id=book_id)
    if request.method == 'POST':
        book.delete()
        return redirect('books_list')
    return render(request, 'compare/delete_book.html', {'book': book})

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def handle_uploaded_file(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        ext = file.name.split('.')[-1].lower()

        if ext not in ['pdf', 'docx', 'doc']:
            return JsonResponse({'error': 'Only PDF, DOCX, and DOC files are allowed.'}, status=400)

        file_content = extract_file_content(file)
        if file_content:
            return JsonResponse({'file_content': file_content})

        return JsonResponse({'error': 'Failed to extract content from the file.'}, status=400)

    return JsonResponse({'error': 'No file uploaded or invalid request.'}, status=400)
